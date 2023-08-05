import pandas as pd
import os
import xlsxwriter
from openpyxl.utils import get_column_letter
import sys

df = pd.read_excel('data\db_2023_07_29.xlsx')
_df_ = df.copy()

COL_REGION = '26.0. Сўнги 3 йилда хизмат кўрсатувчи банкни ўзгартирдингизми? (Қайси банкдан)'



df_columns = pd.read_excel('data\\db_columns.xlsx')
df_columns.dropna(axis=1, how='all', inplace=True)

### drop "others" columns and columns with numeric inputs
COLUMNS = [c for c in df_columns.columns.values.tolist() if not '(Бошқа)' in c]

### drop multiple choice first question
mult_droppers = []

for i, c in enumerate(COLUMNS):
    if i < len(COLUMNS)-1:
        if c in COLUMNS[i+1] and '/' in COLUMNS[i+1]:
            mult_droppers.append(i)

for index in sorted(mult_droppers, reverse=True):
    del COLUMNS[index]
##############################################

COLUMNS.pop(COLUMNS.index('5. ТАРМОҚНИ ТАНЛАНГ:'))
COLUMNS.pop(COLUMNS.index('1. Респондент (корхона) жойлашган ҳудуд:'))


df.dropna(subset=['26.0. Сўнги 3 йилда хизмат кўрсатувчи банкни ўзгартирдингизми? (Қайси банкдан)'], inplace=True)

COLUMNS = ['26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/Хизмат кўрсатиш (комиссия) нархлари паст',
           '26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/Кредит ажратиш қулайлиги',
           '26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/Банк ходимлари малакали',
           '26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/Хизмат кўрсатиш сифати юқори',
           '26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/Банк хизматлари (операциялари) турининг кўплиги',
           '26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/Бошқа']


def crosstab(df, columns):
    ctab = df.groupby(by=[COL_REGION], as_index=True)[columns].sum()
    
    new_cols = [c.replace('26.1. Нима учун ҳозирги хизмат кўрсатувчи банкни танлагансиз?/', '') for c in df.columns.values]
    old_cols = df.columns.values.tolist()
    z_columns = dict(zip(old_cols, new_cols))
    ctab.rename(columns=z_columns, inplace=True)
    df_banks = df[COL_REGION].value_counts()
    # df_merged = pd.merge(ctab, df_banks, how='left', left_index=True, right_index=True)
    
    ctab = ctab.div(df_banks, axis=0).multiply(100).round(1)
    ctab.index.name = COL_REGION
    return ctab


def to_excel(ctabs, filename_out):
    workbook = xlsxwriter.Workbook(filename_out)
    sheet_name = 'Charts'
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.set_column('A:A', 42.71)
    worksheet.set_column('B:AI', 10.43)

    cur_row = 1
    text_format = workbook.add_format({'text_wrap': True})
    tarmoq_format = workbook.add_format({'text_wrap': True,
                                         'bg_color': '#D8E4BC'})
    reg_format = workbook.add_format({'text_wrap': True,
                                      'bg_color': '#B7DEE8'})

    tarmoq_idx = ctabs[0].columns.values.tolist().index('Бошқа')+1
    
    for ctab in ctabs:
        ctab = ctab.reset_index(drop=False).T.reset_index(drop=False).T
        init_row = cur_row
        for ri, index in enumerate(ctab.index.values.tolist()):
            for ci, column in enumerate(ctab.columns.values.tolist()):
                letter = get_column_letter(ci+1)
                value = ctab.at[index, column]
                if ri == 0:
                    if ci <= tarmoq_idx:
                        worksheet.write(f'{letter}{cur_row+1}', value, reg_format)
                    else:
                        worksheet.write(f'{letter}{cur_row+1}', value, tarmoq_format)
                else:
                    worksheet.write(f'{letter}{cur_row+1}', value)
            cur_row += 1


        chart_reg = workbook.add_chart({'type': 'column', 
                                        'subtype': 'percent_stacked'})
        


        tarmoq_letter_start = get_column_letter(tarmoq_idx+2)
        reg_letter_end = get_column_letter(tarmoq_idx+1)

        CHART_WIDTH = 800
        CHART_HEIGHT = 300
        # for i in range(init_row+2, cur_row+1):
        #     letter = get_column_letter(i)
        #     chart_reg.add_series({
        #             'name':f'={sheet_name}!$A${i}',
        #             'categories': f'={sheet_name}!$B${init_row+1}:${reg_letter_end}${init_row+1}',
        #             'values':     f'={sheet_name}!B{i}:{reg_letter_end}{i}',
        #             'data_labels': {
        #                             'value': True,
        #                             'font': {'size': 10}
        #                             }
        #         })
            
        # chart_reg.set_size({'width': CHART_WIDTH, 'height': CHART_HEIGHT})
        # chart_reg.set_x_axis({'num_font':  {'name': 'Arial', 'size': 8}})
        # chart_reg.set_title({
        #                 'name': f'={sheet_name}!$A${init_row+1}',
        #                 'name_font': {
        #                     'name': 'Arial',
        #                     'color': '#808080',
        #                     'size': 10
        #                 },
        #             })

        # worksheet.insert_chart(f'A{cur_row+1}', chart_reg)

        for i in range(2, tarmoq_idx+2):
            letter = get_column_letter(i)
            print(init_row, cur_row, i, )
            print(f'={sheet_name}!B{init_row+2}:{reg_letter_end}{cur_row+1}')
            chart_reg.add_series({
                    'name':f'={sheet_name}!${letter}${init_row+1}',
                    'categories': f'={sheet_name}!$A${init_row+2}:$A${cur_row+1}',
                    'values':     f'={sheet_name}!{letter}{init_row+2}:{letter}{cur_row+1}',
                    'data_labels': {
                                    'value': True,
                                    'font': {'size': 10}
                                    }
                })
            
        chart_reg.set_size({'width': CHART_WIDTH, 'height': CHART_HEIGHT})
        chart_reg.set_x_axis({'num_font':  {'name': 'Arial', 'size': 8}})
        chart_reg.set_title({
                        'name': f'={sheet_name}!$A${init_row+1}',
                        'name_font': {
                            'name': 'Arial',
                            'color': '#808080',
                            'size': 10
                        },
                    })

        worksheet.insert_chart(f'A{cur_row+1}', chart_reg)


        cur_row += 15

        
        
    
    workbook.close()


ctabs = []

ctabs.append(crosstab(df, COLUMNS))

to_excel(ctabs, 'out\\result.xlsx')