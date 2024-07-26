import pandas as pd
import os
import xlsxwriter
from openpyxl.utils import get_column_letter
import sys
import unicodedata

COL_REGION = 'region'
COL_TARMOQ = '5. Корхонанинг асосий иқтисодий фаолият тури:'

def get_df():


    def contains_cyrillic(input_string):
        for char in input_string:
            if 'CYRILLIC' in unicodedata.name(char, ''):
                return True
        return False
    

    df = pd.read_excel('data\db_2024_07_26.xlsx')

    df[COL_TARMOQ] = df[COL_TARMOQ].mask(df[COL_TARMOQ] == 'Хизмат кўрсатиш', df['5.2. Хизмат кўрсатиш соҳалари:'])
    df[COL_TARMOQ] = df[COL_TARMOQ].mask(df[COL_TARMOQ] == 'Саноат', df['5.1. Саноат фаолият турлари:'])
    # df['5.1. Саноат фаолият турлари:'] = df['5.1. Саноат фаолият турлари:'].fillna(df['5. Корхонанинг асосий иқтисодий фаолият тури:'])
    # df['5.1. Саноат фаолият турлари:'] = df['5.1. Саноат фаолият турлари:'].fillna(df['5.2. Хизмат кўрсатиш соҳалари:'])
    df_columns = df.copy()

    

    COLUMNS = ['2-9. Фаолиятингиз давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини қандай баҳолайсиз?/Солиқ инспекцияси',
            '1-22. Солиқ тўлаш жараёни бўйича тажрибангизга таяниб, қуйидаги фикрларга муносабатингизни билдиринг?/Солиқ ходимлари профессионал ва холис']

    df_columns.dropna(axis=1, how='all', inplace=True)

    ### drop "others" columns and columns with numeric inputs
    COLUMNS = [c for c in df_columns.columns.values.tolist() if not '(Бошқа)' in c and (contains_cyrillic(c) or c in ['region', 'district'])]

    ### drop multiple choice first question
    mult_droppers = []

    for i, c in enumerate(COLUMNS):
        if i < len(COLUMNS)-1:
            if c in COLUMNS[i+1] and '/' in COLUMNS[i+1]:
                mult_droppers.append(i)

    for index in sorted(mult_droppers, reverse=True):
        del COLUMNS[index]
    ##############################################
    COLUMNS.pop(COLUMNS.index(COL_REGION))
    COLUMNS.pop(COLUMNS.index(COL_TARMOQ))
    COLUMNS.pop(COLUMNS.index('district'))
    return df, COLUMNS

def crosstab(df, column):
    ctab_reg = pd.crosstab(index=df[column], columns=[df[COL_REGION]], normalize='columns')
    ctab_reg = ctab_reg.multiply(100).round(1)
    ctab_tar = pd.crosstab(index=df[column], columns=[df[COL_TARMOQ]], normalize='columns', margins=True)
    ctab_tar = ctab_tar.multiply(100).round(1)
    ctab = pd.merge(ctab_reg, ctab_tar, how='left', left_index=True, right_index=True)
    return ctab


def to_excel(ctabs, filename_out):
    workbook = xlsxwriter.Workbook(filename_out)
    sheet_name = 'Charts'
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.set_column('A:A', 42.71)
    worksheet.set_column('B:AI', 10.43)

    cur_row = 1
    text_format = workbook.add_format({'text_wrap': True})
    tarmoq_format = workbook.add_format({'text_wrap': True,
                                         'bg_color': '#D8E4BC'})
    reg_format = workbook.add_format({'text_wrap': True,
                                      'bg_color': '#B7DEE8'})
    


  

    for ctab in ctabs:
        ctab = ctab.reset_index(drop=False).T.reset_index(drop=False).T
        init_row = cur_row

        for idx, c in enumerate(ctab.head(1).values.tolist()[0]):
            if idx == 0:
                continue
            tarmoq_idx = idx
            if not(c.strip().endswith('вилояти') or c.strip().endswith('шаҳри')):
                break
            
        for ri, index in enumerate(ctab.index.values.tolist()):
            for ci, column in enumerate(ctab.columns.values.tolist()):
                letter = get_column_letter(ci+1)
                value = ctab.at[index, column]
                if ri == 0:
                    if ci < tarmoq_idx:
                        worksheet.write(f'{letter}{cur_row+1}', value, reg_format)
                    else:
                        worksheet.write(f'{letter}{cur_row+1}', value, tarmoq_format)
                else:
                    worksheet.write(f'{letter}{cur_row+1}', value)
            cur_row += 1


        chart_reg = workbook.add_chart({'type': 'column', 
                                        'subtype': 'percent_stacked'})
        


        

        CHART_WIDTH = 800
        CHART_HEIGHT = 400


        #### find where regions stop and tarmoq start
        for idx, c in enumerate(ctab.head(1).values.tolist()[0]):
            if idx == 0:
                continue
            if not (c.strip().endswith('вилояти') or c.strip().endswith('шаҳри')):
                break
            reg_letter_end = get_column_letter(idx+1)
        #############################################################
            
        for i in range(init_row+2, cur_row+1):
            letter = get_column_letter(i)
            chart_reg.add_series({
                    'name':f'={sheet_name}!$A${i}',
                    'categories': f'={sheet_name}!$B${init_row+1}:${reg_letter_end}${init_row+1}',
                    'values':     f'={sheet_name}!B{i}:{reg_letter_end}{i}',
                    'data_labels': {
                                    'value': True,
                                    'font': {'size': 10}
                                    }
                })
            
        chart_reg.set_size({'width': CHART_WIDTH, 'height': CHART_HEIGHT})
        chart_reg.set_x_axis({'num_font':  {'name': 'Arial', 'size': 8}})
        chart_reg.set_title({
                        'name': f'={sheet_name}!$A${init_row+1}',
                        'name_font': {
                            'name': 'Arial',
                            'color': '#808080',
                            'size': 10
                        },
                    })

        worksheet.insert_chart(f'A{cur_row+1}', chart_reg)


        chart_tarmoq = workbook.add_chart({'type': 'column', 
                                    'subtype': 'percent_stacked'})

        tarmoq_letter_end = get_column_letter(ctab.shape[1])
        for idx, c in enumerate(ctab.head(1).values.tolist()[0]):
            if idx == 0:
                continue
            tarmoq_letter_start = get_column_letter(idx+1)
            if not(c.strip().endswith('вилояти') or c.strip().endswith('шаҳри')):
                break

        for i in range(init_row+2, cur_row+1):
            letter = get_column_letter(i)
            chart_tarmoq.add_series({
                    'name':f'={sheet_name}!$A${i}',
                    'categories': f'={sheet_name}!${tarmoq_letter_start}${init_row+1}:${tarmoq_letter_end}${init_row+1}',
                    'values':     f'={sheet_name}!{tarmoq_letter_start}{i}:{tarmoq_letter_end}{i}',
                    'data_labels': {
                                    'value': True,
                                    'font': {'size': 10}
                                    }
                })
            
        chart_tarmoq.set_size({'width': CHART_WIDTH, 'height': CHART_HEIGHT})
        chart_tarmoq.set_x_axis({'num_font':  {'name': 'Arial', 'size': 8}})
        chart_tarmoq.set_title({
                        'name': f'={sheet_name}!$A${init_row+1}',
                        'name_font': {
                            'name': 'Arial',
                            'color': '#808080',
                            'size': 10
                        },
                    })

        worksheet.insert_chart(f'{"O"}{cur_row+1}', chart_tarmoq)

        cur_row += 22

        
        
    
    workbook.close()

df, COLUMNS = get_df()

ctabs = []
for column in COLUMNS:
    ctabs.append(crosstab(df, column))


to_excel(ctabs, 'out\\result.xlsx')